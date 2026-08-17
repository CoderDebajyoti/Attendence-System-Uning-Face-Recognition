# ==============================================================================
# Face Recognition Attendance System - Attendance Report Service
# ==============================================================================

import csv
import logging
import os
from pathlib import Path
from datetime import datetime
from src.core.config import ConfigLoader
from src.core.models import Attendance
from src.repositories.attendance_repository import AttendanceRepository
from src.services.attendance_analytics_service import AttendanceAnalyticsService
from src.utils.time_helper import get_current_date, get_local_now

logger = logging.getLogger("app.reports")

class AttendanceReportService:
    """
    Service responsible for validating report configurations, retrieving filtered datasets,
    rendering previews, and exporting CSV/Excel spreadsheet reports.
    """
    _instance = None

    @classmethod
    def get_instance(cls, settings=None):
        if cls._instance is None:
            cls._instance = cls(settings)
        return cls._instance

    def __init__(self, settings=None) -> None:
        self.settings = settings or ConfigLoader.load_config()
        self.repo = AttendanceRepository()
        self.analytics_service = AttendanceAnalyticsService.get_instance()
        logger.info("AttendanceReportService initialized successfully.")

    def validate_date_range(self, start_date_str: str, end_date_str: str) -> tuple[bool, str]:
        """
        Ensures date strings are non-empty, match YYYY-MM-DD format, and start_date <= end_date.
        """
        if not start_date_str or not end_date_str:
            return False, "Start Date and End Date are required."
        try:
            start_dt = datetime.strptime(start_date_str, "%Y-%m-%d")
            end_dt = datetime.strptime(end_date_str, "%Y-%m-%d")
            if start_dt > end_dt:
                return False, "Start Date cannot be after End Date."
            return True, "Valid date range."
        except ValueError:
            return False, "Dates must follow the YYYY-MM-DD format."

    def generate_report_data(
        self,
        start_date: str,
        end_date: str,
        status: str | None = None,
        department_id: int | None = None,
        course_id: int | None = None,
        source: str | None = None,
        student_id: int | None = None,
        search_query: str | None = None
    ) -> dict:
        """
        Fetches filtered logs and returns a summary payload.
        """
        # 1. Fetch filtered logs list
        records = self.repo.list_attendance(
            search_query=search_query,
            start_date=start_date,
            end_date=end_date,
            status=status,
            department_id=department_id,
            course_id=course_id,
            source=source
        )
        
        # Additional client-side filter for student_id if supplied
        if student_id:
            records = [r for r in records if r.student_id == student_id]

        # 2. Get date range statistics
        stats = self.analytics_service.get_date_range_statistics(
            start_date=start_date,
            end_date=end_date,
            department_id=department_id,
            course_id=course_id,
            student_id=student_id
        )

        return {
            "records": records,
            "summary": stats,
            "start_date": start_date,
            "end_date": end_date
        }

    def generate_safe_export_path(self, file_extension: str, start_date: str, end_date: str) -> Path:
        """
        Generates a standardized non-colliding filename inside the configured export folder.
        """
        export_dir = Path(self.settings.export_path)
        os.makedirs(export_dir, exist_ok=True)
        
        timestamp = get_local_now().strftime("%Y%m%d_%H%M%S")
        filename = f"attendance_report_{start_date}_to_{end_date}_{timestamp}.{file_extension}"
        return export_dir / filename

    def export_to_csv(self, report_data: dict) -> tuple[bool, str]:
        """
        Exports currently filtered attendance records to a CSV file.
        """
        records = report_data.get("records", [])
        if not records:
            return False, "No attendance records found for the selected filters."

        start_date = report_data["start_date"]
        end_date = report_data["end_date"]
        csv_path = self.generate_safe_export_path("csv", start_date, end_date)

        try:
            with open(csv_path, mode="w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                # Header row
                writer.writerow([
                    "Attendance ID", "Student ID", "Student Name", "Roll Number",
                    "Department", "Course", "Date", "Time In", "Status", "Source", "Recognition Score"
                ])
                # Data rows
                for r in records:
                    dept = r.student.department.code if r.student.department else "-"
                    course = r.student.course.code if r.student.course else "-"
                    score_val = f"{r.recognition_score:.2f}" if r.recognition_score is not None else "-"
                    writer.writerow([
                        r.id,
                        r.student.student_code,
                        f"{r.student.first_name} {r.student.last_name}",
                        r.student.roll_number,
                        dept,
                        course,
                        r.date,
                        r.time_in,
                        r.status,
                        r.source,
                        score_val
                    ])
            
            logger.info(f"Successfully generated CSV report at: {csv_path}")
            return True, f"Report exported successfully to:\n{csv_path.name}"
        except Exception as e:
            logger.error(f"Failed to write CSV report: {e}", exc_info=True)
            return False, f"Export failed: {e}"

    def export_to_excel(self, report_data: dict) -> tuple[bool, str]:
        """
        Exports reports into a multi-sheet Excel spreadsheet using openpyxl.
        """
        records = report_data.get("records", [])
        if not records:
            return False, "No attendance records found for the selected filters."

        # Verify dependency is loaded
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            return False, "Excel library (openpyxl) is not installed."

        start_date = report_data["start_date"]
        end_date = report_data["end_date"]
        summary = report_data["summary"]
        excel_path = self.generate_safe_export_path("xlsx", start_date, end_date)

        try:
            wb = openpyxl.Workbook()
            
            # Sheet 1: Records View
            ws_records = wb.active
            ws_records.title = "Attendance Records"
            
            # Styling headers
            header_fill = PatternFill(start_color="2A3F54", end_color="2A3F54", fill_type="solid")
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            align_center = Alignment(horizontal="center", vertical="center")
            
            headers = [
                "Attendance ID", "Student ID", "Student Name", "Roll Number",
                "Department", "Course", "Date", "Time In", "Status", "Source", "Recognition Score"
            ]
            ws_records.append(headers)
            
            for col_idx in range(1, len(headers) + 1):
                cell = ws_records.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = align_center

            # Add record rows
            for idx, r in enumerate(records):
                dept = r.student.department.code if r.student.department else "-"
                course = r.student.course.code if r.student.course else "-"
                score_val = r.recognition_score if r.recognition_score is not None else "-"
                
                row_data = [
                    r.id,
                    r.student.student_code,
                    f"{r.student.first_name} {r.student.last_name}",
                    r.student.roll_number,
                    dept,
                    course,
                    r.date,
                    r.time_in,
                    r.status,
                    r.source,
                    score_val
                ]
                ws_records.append(row_data)

            # Autofit column widths
            for col in ws_records.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws_records.column_dimensions[col_letter].width = max(max_len + 3, 12)

            # Sheet 2: Summary Stats
            ws_summary = wb.create_sheet(title="Summary Report")
            
            # Title banner
            ws_summary.merge_cells("A1:C1")
            ws_summary["A1"] = "Attendance Report Summary"
            ws_summary["A1"].font = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
            ws_summary["A1"].fill = PatternFill(start_color="1ABB9C", end_color="1ABB9C", fill_type="solid")
            ws_summary["A1"].alignment = align_center
            ws_summary.row_dimensions[1].height = 40

            summary_rows = [
                ("Start Date", start_date),
                ("End Date", end_date),
                ("Total Records", summary["total_records"]),
                ("Present", summary["present"]),
                ("Late", summary["late"]),
                ("Absent", summary["absent"]),
                ("Excused", summary["excused"]),
                ("Total Opportunities", summary["opportunities"]),
                ("Overall Attendance Rate", f"{summary['rate']}%")
            ]
            
            for idx, (label, val) in enumerate(summary_rows):
                row_num = idx + 3
                ws_summary.cell(row=row_num, column=1, value=label).font = Font(name="Segoe UI", bold=True)
                ws_summary.cell(row=row_num, column=2, value=val).font = Font(name="Segoe UI")
            
            ws_summary.column_dimensions["A"].width = 24
            ws_summary.column_dimensions["B"].width = 16

            wb.save(excel_path)
            logger.info(f"Successfully generated Excel report at: {excel_path}")
            return True, f"Report exported successfully to:\n{excel_path.name}"
        except Exception as e:
            logger.error(f"Failed to write Excel report: {e}", exc_info=True)
            return False, f"Export failed: {e}"
